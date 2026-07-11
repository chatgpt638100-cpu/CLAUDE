"""
Excel Attendance Export Module
Automatically exports attendance to Excel file
"""
import os
from datetime import datetime
try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    print("Warning: openpyxl not installed. Excel export will not work.")
    print("Install with: pip install openpyxl")


def export_attendance_to_excel(attendance_data, output_dir='data/attendance_excel'):
    """
    Export attendance to Excel file
    
    Args:
        attendance_data: List of attendance records
        output_dir: Directory to save Excel files
    """
    if not EXCEL_AVAILABLE:
        print("Excel export unavailable - openpyxl not installed")
        return None
    
    # Create output directory
    os.makedirs(output_dir, exist_ok=True)
    
    # Generate filename with current date
    date_str = datetime.now().strftime("%Y-%m-%d")
    filename = os.path.join(output_dir, f"Attendance_{date_str}.xlsx")
    
    # Check if file exists, if so load it, otherwise create new
    if os.path.exists(filename):
        wb = openpyxl.load_workbook(filename)
        ws = wb.active
    else:
        # Create new workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Attendance"
        
        # Create header row
        headers = ['Date', 'Time', 'Student Name', 'Status', 'Confidence']
        ws.append(headers)
        
        # Style header row
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
    
    # Add attendance data
    for record in attendance_data:
        date_time = datetime.strptime(record['timestamp'], "%Y-%m-%d %H:%M:%S")
        row = [
            date_time.strftime("%Y-%m-%d"),
            date_time.strftime("%H:%M:%S"),
            record['student_name'],
            record['status'],
            f"{record['confidence']:.2f}"
        ]
        ws.append(row)
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save workbook
    wb.save(filename)
    print(f"✓ Attendance exported to Excel: {filename}")
    
    return filename


def mark_attendance_to_excel(student_name, confidence):
    """
    Mark single student attendance and export to Excel
    
    Args:
        student_name: Name of student
        confidence: Recognition confidence
    """
    if not EXCEL_AVAILABLE:
        return False
    
    # Create attendance record
    attendance_record = {
        'student_name': student_name,
        'timestamp': datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        'status': 'present',
        'confidence': confidence
    }
    
    # Export to Excel
    export_attendance_to_excel([attendance_record])
    
    return True


if __name__ == "__main__":
    # Test Excel export
    print("Testing Excel attendance export...")
    
    test_data = [
        {
            'student_name': 'Bhava',
            'timestamp': datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            'status': 'present',
            'confidence': 0.89
        },
        {
            'student_name': 'Vishal',
            'timestamp': datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            'status': 'present',
            'confidence': 0.92
        }
    ]
    
    export_attendance_to_excel(test_data)
    print("Test complete!")
